"""Leads — the core of the tool."""
import re
from datetime import date

from flask import Blueprint, flash, redirect, render_template, request, session, url_for

from . import db
from .db import execute, log_activity, query_all, query_one
from .security import admin_required, login_required, trainer_required
from .utils import Pagination
from .utils.vk_validator import VKValidator

bp = Blueprint("leads", __name__)

STATUSES = ["new", "contacted", "replied", "joined_channel", "participated", "returning",
           "declined", "unresponsive"]
STATUS_LABELS = {
    "new": "Новый", "contacted": "Написали", "replied": "Ответил",
    "joined_channel": "Вступил в канал", "participated": "Участвовал",
    "returning": "Возвращается", "declined": "Отказался", "unresponsive": "Не отвечает",
}
ROUND_SLOTS = ["12:00", "18:00", "00:00"]


def normalize_vk_id(raw: str) -> str:
    raw = (raw or "").strip().lower()
    raw = re.sub(r"^https?://", "", raw)
    raw = re.sub(r"^(www\.|m\.)?vk\.(com|ru)/", "", raw)
    raw = raw.lstrip("@").rstrip("/")
    raw = raw.split("?")[0]
    return raw


def next_manager_id():
    row = query_one(
        "SELECT m.id FROM managers m "
        "LEFT JOIN leads l ON l.assigned_manager_id = m.id "
        "AND l.status NOT IN ('declined', 'unresponsive') "
        "WHERE m.role='manager' AND m.is_active=1 "
        "GROUP BY m.id ORDER BY COUNT(l.id) ASC, m.id ASC LIMIT 1")
    return row["id"] if row else None


def add_lead(vk_url, source_group, found_by_manager_id, name="", assign_to_finder=True):
    # Расширенная валидация VK ссылки
    is_valid, message = VKValidator.is_valid_vk_url(vk_url, check_exists=False)
    if not is_valid:
        return None, False, f"❌ {message}"
    
    vk_id = VKValidator.extract_id(vk_url)
    if not vk_id:
        return None, False, "❌ Не удалось распознать VK ID"

    # Проверка на дубликат — без учёта регистра, иначе "Durov" и "durov"
    # проходят как разные лиды (раньше это ловил только риск-модуль
    # постфактум; тут — превентивно, до записи в базу).
    existing = query_one("SELECT id FROM leads WHERE LOWER(vk_id)=LOWER(?)", (vk_id,))
    if existing:
        return existing["id"], False, "⚠️ Лид уже существует"

    manager_id = found_by_manager_id if assign_to_finder else next_manager_id()
    lead_id = execute(
        "INSERT INTO leads (vk_id, vk_url, name, source_group, assigned_manager_id) "
        "VALUES (?, ?, ?, ?, ?)",
        (vk_id, vk_url.strip(), name.strip(), source_group.strip(), manager_id))
    return lead_id, True, "✅ Лид добавлен"


def bulk_import(lines, source_group, assigned_manager_id=None):
    """Массовый импорт с указанием менеджера"""
    added = 0
    duplicates = 0
    skipped = 0
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Валидация перед импортом
        is_valid, _ = VKValidator.is_valid_vk_url(line, check_exists=False)
        if not is_valid:
            skipped += 1
            continue
        
        # Проверяем существование — без учёта регистра (см. add_lead)
        vk_id = VKValidator.extract_id(line)
        existing = query_one("SELECT id FROM leads WHERE LOWER(vk_id)=LOWER(?)", (vk_id,))
        if existing:
            duplicates += 1
            continue
        
        # Если менеджер указан - назначаем его, иначе - round-robin
        manager_id = assigned_manager_id if assigned_manager_id else next_manager_id()
        
        execute(
            "INSERT INTO leads (vk_id, vk_url, name, source_group, assigned_manager_id) "
            "VALUES (?, ?, ?, ?, ?)",
            (vk_id, line.strip(), "", source_group.strip(), manager_id))
        added += 1
    
    return added, duplicates, skipped


def vk_chat_url(vk_id: str) -> str:
    return f"https://vk.ru/im?sel={vk_id}"


def _can_touch(lead, manager_id, role):
    return role == "admin" or lead["assigned_manager_id"] == manager_id


@bp.route("/leads")
@trainer_required
def index():
    manager_id = session["manager_id"]
    role = session["role"]
    status_filter = request.args.get("status", "")
    tag_filter = request.args.get("tag", "").strip()
    search = request.args.get("q", "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)

    if role == "admin":
        base = "SELECT l.*, m.name manager_name FROM leads l LEFT JOIN managers m ON m.id=l.assigned_manager_id"
        where, params = [], []
    else:
        base = "SELECT l.*, m.name manager_name FROM leads l LEFT JOIN managers m ON m.id=l.assigned_manager_id"
        where, params = ["l.assigned_manager_id=?"], [manager_id]

    if status_filter:
        where.append("l.status=?")
        params.append(status_filter)

    if tag_filter:
        where.append("l.tags LIKE ?")
        params.append(f'%"{tag_filter}"%')

    if search:
        where.append("(l.name LIKE ? OR l.vk_id LIKE ? OR l.game_id LIKE ?)")
        like = f"%{search}%"
        params.extend([like, like, like])

    sql = base + (" WHERE " + " AND ".join(where) if where else "") + " ORDER BY l.found_at DESC LIMIT 200"
    all_leads = query_all(sql, tuple(params))
    
    pagination = Pagination(all_leads, page, per_page)
    scripts = query_all("SELECT id, label FROM scripts WHERE is_active=1 ORDER BY label")

    # Заметки только для лидов на текущей странице (не для всех 200) —
    # избегаем N+1: один запрос с IN (...) вместо запроса на каждый лид.
    page_leads = [dict(r) for r in pagination.current_items]
    page_ids = [l["id"] for l in page_leads]
    notes_by_lead = {}
    if page_ids:
        placeholders = ",".join("?" for _ in page_ids)
        notes = query_all(
            f"SELECT n.*, m.name manager_name FROM lead_notes n "
            f"LEFT JOIN managers m ON m.id=n.manager_id "
            f"WHERE n.lead_id IN ({placeholders}) ORDER BY n.created_at DESC", page_ids)
        for n in notes:
            notes_by_lead.setdefault(n["lead_id"], []).append(dict(n))
    for l in page_leads:
        l["notes_list"] = notes_by_lead.get(l["id"], [])

    check_result = session.pop("lead_check_result", None)

    return render_template("leads.html", 
                          leads=page_leads,
                          pagination=pagination,
                          scripts=[dict(s) for s in scripts],
                          statuses=STATUSES, 
                          status_labels=STATUS_LABELS, 
                          status_filter=status_filter,
                          tag_filter=tag_filter,
                          search=search,
                          is_admin=(role == "admin"),
                          check_result=check_result)

@bp.route("/leads/export.csv")
@trainer_required
def export_csv():
    manager_id = session["manager_id"]
    role = session["role"]
    status_filter = request.args.get("status", "")
    search = request.args.get("q", "").strip()

    base = "SELECT l.*, m.name manager_name FROM leads l LEFT JOIN managers m ON m.id=l.assigned_manager_id"
    if role == "admin":
        where, params = [], []
    else:
        where, params = ["l.assigned_manager_id=?"], [manager_id]
    if status_filter:
        where.append("l.status=?")
        params.append(status_filter)
    if search:
        where.append("(l.name LIKE ? OR l.vk_id LIKE ? OR l.game_id LIKE ?)")
        like = f"%{search}%"
        params.extend([like, like, like])
    sql = base + (" WHERE " + " AND ".join(where) if where else "") + " ORDER BY l.found_at DESC"
    rows = query_all(sql, tuple(params))

    import csv
    import io
    from flask import Response
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["ID", "Имя", "VK ID", "VK ссылка", "Статус", "Источник", "Менеджер",
                      "Баланс", "Game ID", "Найден", "Первый контакт"])
    for r in rows:
        writer.writerow([
            r["id"], r["name"], r["vk_id"], r["vk_url"], STATUS_LABELS.get(r["status"], r["status"]),
            r["source_group"], r["manager_name"] or "", r["balance"], r["game_id"] or "",
            r["found_at"], r["first_contacted_at"] or "",
        ])
    csv_data = "\ufeff" + buf.getvalue()  # BOM — чтобы Excel сразу видел кириллицу как UTF-8
    return Response(csv_data, mimetype="text/csv", headers={
        "Content-Disposition": "attachment; filename=leads_export.csv"
    })


@bp.route("/leads/export.xlsx")
@trainer_required
def export_xlsx():
    try:
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        flash("Для экспорта в Excel нужен пакет openpyxl (pip install openpyxl). Пока доступен CSV.", "error")
        return redirect(url_for("leads.index"))

    manager_id = session["manager_id"]
    role = session["role"]
    status_filter = request.args.get("status", "")
    search = request.args.get("q", "").strip()

    base = "SELECT l.*, m.name manager_name FROM leads l LEFT JOIN managers m ON m.id=l.assigned_manager_id"
    if role == "admin":
        where, params = [], []
    else:
        where, params = ["l.assigned_manager_id=?"], [manager_id]
    if status_filter:
        where.append("l.status=?")
        params.append(status_filter)
    if search:
        where.append("(l.name LIKE ? OR l.vk_id LIKE ? OR l.game_id LIKE ?)")
        like = f"%{search}%"
        params.extend([like, like, like])
    sql = base + (" WHERE " + " AND ".join(where) if where else "") + " ORDER BY l.found_at DESC"
    rows = query_all(sql, tuple(params))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лиды"
    headers = ["ID", "Имя", "VK ID", "VK ссылка", "Статус", "Источник", "Менеджер", "Баланс", "Game ID", "Найден"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2EC5FF")
    for r in rows:
        ws.append([r["id"], r["name"], r["vk_id"], r["vk_url"], STATUS_LABELS.get(r["status"], r["status"]),
                   r["source_group"], r["manager_name"] or "", r["balance"], r["game_id"] or "", r["found_at"]])
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import send_file
    return send_file(buf, as_attachment=True, download_name="leads_export.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/leads/check", methods=["POST"])
@trainer_required
def check():
    """Проверка ссылки VK перед добавлением: показывает, существует ли уже
    такой лид, и если да — кто именно его ведёт, чтобы менеджер не писал
    повторно тому, кого уже ведёт другой менеджер."""
    raw = request.form.get("vk_url", "").strip()

    if not raw:
        flash("❌ Укажите ссылку на профиль ВК.", "error")
        return redirect(url_for("leads.index"))

    is_valid, message = VKValidator.is_valid_vk_url(raw, check_exists=False)
    if not is_valid:
        flash(f"❌ {message}", "error")
        return redirect(url_for("leads.index"))

    vk_id = VKValidator.extract_id(raw)
    if not vk_id:
        flash("❌ Не удалось распознать VK ID.", "error")
        return redirect(url_for("leads.index"))

    existing = query_one(
        "SELECT l.id, l.name, l.status, l.assigned_manager_id, m.name as manager_name "
        "FROM leads l LEFT JOIN managers m ON m.id = l.assigned_manager_id "
        "WHERE l.vk_id=?", (vk_id,))

    session["lead_check_result"] = {
        "vk_id": vk_id,
        "canonical": f"vk.ru/{vk_id}",
        "exists": bool(existing),
        "existing_name": existing["name"] if existing else None,
        "owner_name": existing["manager_name"] if existing else None,
        "owner_is_me": bool(existing and existing["assigned_manager_id"] == session["manager_id"]),
        "status_label": STATUS_LABELS.get(existing["status"], existing["status"]) if existing else None,
    }
    return redirect(url_for("leads.index"))


@bp.route("/leads/add", methods=["POST"])
@trainer_required
def add():
    manager_id = session["manager_id"]
    vk_url = request.form.get("vk_url", "").strip()
    source_group = request.form.get("source_group", "").strip()
    name = request.form.get("name", "").strip()
    
    if not vk_url:
        flash("❌ Укажите ссылку на профиль ВК.", "error")
        return redirect(url_for("leads.index"))
    
    is_valid, message = VKValidator.is_valid_vk_url(vk_url, check_exists=False)
    if not is_valid:
        flash(f"❌ {message}", "error")
        return redirect(url_for("leads.index"))

    lead_id, created, msg = add_lead(vk_url, source_group, manager_id, name)
    
    if lead_id is None:
        flash(f"{msg}", "error")
    elif created:
        flash(f"{msg}, назначен вам.", "success")
    else:
        owner = query_one(
            "SELECT m.name, l.status FROM leads l LEFT JOIN managers m ON m.id=l.assigned_manager_id "
            "WHERE l.id=?", (lead_id,))
        flash(f"{msg} — ведёт {owner['name'] if owner else 'неизвестно'} "
             f"(статус: {STATUS_LABELS.get(owner['status'], owner['status'])}). Не пишите повторно.", "warning")
    return redirect(url_for("leads.index"))


@bp.route("/leads/bulk-import", methods=["POST"])
@admin_required
def bulk_import_route():
    """Массовый импорт лидов с выбором менеджера"""
    text = request.form.get("profiles", "")
    source_group = request.form.get("source_group", "").strip()
    manager_id = request.form.get("manager_id")
    
    if not text.strip():
        flash("❌ Введите список ссылок для импорта.", "error")
        return redirect(url_for("leads.index"))
    
    lines = text.splitlines()
    
    # Если менеджер не выбран - используем round-robin
    assigned_manager = None
    manager_name = "автоматически (round-robin)"
    
    if manager_id and manager_id != "":
        manager = query_one("SELECT id, name FROM managers WHERE id=? AND role='manager'", (manager_id,))
        if manager:
            assigned_manager = manager["id"]
            manager_name = manager["name"]
    
    added, duplicates, skipped = bulk_import(lines, source_group, assigned_manager)
    
    flash(f"✅ Добавлено: {added}. Уже были: {duplicates}. Пропущено (неверный формат): {skipped}.", "success")
    flash(f"📌 Лиды назначены: {manager_name}", "info")
    log_activity(f"Админ импортировал {added} лидов, назначены: {manager_name}")
    
    return redirect(url_for("leads.index"))


# Остальные маршруты
@bp.route("/leads/<int:lead_id>/contact", methods=["POST"])
@trainer_required
def mark_contacted(lead_id):
    manager_id, role = session["manager_id"], session["role"]
    lead = query_one("SELECT * FROM leads WHERE id=?", (lead_id,))
    if not lead or not _can_touch(lead, manager_id, role):
        flash("Лид не найден или не ваш.", "error")
        return redirect(url_for("leads.index"))

    script_id = request.form.get("script_id") or None
    execute("INSERT INTO outreach_log (lead_id, manager_id, script_id) VALUES (?, ?, ?)",
            (lead_id, manager_id, script_id))
    if lead["status"] == "new":
        execute("UPDATE leads SET status='contacted', first_contacted_at=datetime('now'), "
                "last_status_change=datetime('now') WHERE id=?", (lead_id,))
    flash("Отмечено: написали.", "success")
    return redirect(url_for("leads.index"))


@bp.route("/leads/<int:lead_id>/status", methods=["POST"])
@trainer_required
def update_status(lead_id):
    manager_id, role = session["manager_id"], session["role"]
    lead = query_one("SELECT * FROM leads WHERE id=?", (lead_id,))
    if not lead or not _can_touch(lead, manager_id, role):
        flash("Лид не найден или не ваш.", "error")
        return redirect(url_for("leads.index"))

    new_status = request.form.get("status", "")
    if new_status not in STATUSES:
        flash("Некорректный статус.", "error")
        return redirect(url_for("leads.index"))
    execute("UPDATE leads SET status=?, last_status_change=datetime('now') WHERE id=?",
            (new_status, lead_id))
    execute("INSERT INTO lead_status_history (lead_id, status, manager_id) VALUES (?, ?, ?)",
            (lead_id, new_status, manager_id))
    flash("Статус обновлён.", "success")
    return redirect(url_for("leads.index"))


@bp.route("/leads/bulk-action", methods=["POST"])
@trainer_required
def bulk_action():
    """Массовые действия над УЖЕ существующими лидами (не путать с
    /leads/bulk-import, который добавляет НОВЫХ). Менеджер может массово
    менять статус только своим лидам; переназначение — только админ."""
    manager_id, role = session["manager_id"], session["role"]
    try:
        lead_ids = [int(x) for x in request.form.getlist("lead_ids")]
    except ValueError:
        lead_ids = []
    action = request.form.get("action", "")

    if not lead_ids:
        flash("Не выбрано ни одного лида.", "error")
        return redirect(url_for("leads.index"))

    if role != "admin":
        placeholders = ",".join("?" for _ in lead_ids)
        owned = query_all(
            f"SELECT id FROM leads WHERE id IN ({placeholders}) AND assigned_manager_id=?",
            lead_ids + [manager_id])
        lead_ids = [r["id"] for r in owned]
        if not lead_ids:
            flash("Ни один из выбранных лидов не назначен вам.", "error")
            return redirect(url_for("leads.index"))

    placeholders = ",".join("?" for _ in lead_ids)

    if action.startswith("status:"):
        new_status = action.split(":", 1)[1]
        if new_status not in STATUSES:
            flash("Некорректный статус.", "error")
            return redirect(url_for("leads.index"))
        execute(f"UPDATE leads SET status=?, last_status_change=datetime('now') WHERE id IN ({placeholders})",
                [new_status] + lead_ids)
        for lid in lead_ids:
            execute("INSERT INTO lead_status_history (lead_id, status, manager_id) VALUES (?, ?, ?)",
                    (lid, new_status, manager_id))
        log_activity(f"Массовая смена статуса на «{STATUS_LABELS.get(new_status, new_status)}» "
                     f"для {len(lead_ids)} лидов.", manager_id)
        flash(f"✅ Статус обновлён у {len(lead_ids)} лидов.", "success")

    elif action.startswith("reassign:"):
        if role != "admin":
            flash("Переназначать лидов может только админ.", "error")
            return redirect(url_for("leads.index"))
        try:
            new_manager_id = int(action.split(":", 1)[1])
        except ValueError:
            flash("Некорректный менеджер.", "error")
            return redirect(url_for("leads.index"))
        target = query_one("SELECT id, name FROM managers WHERE id=? AND role='manager'", (new_manager_id,))
        if not target:
            flash("Менеджер не найден.", "error")
            return redirect(url_for("leads.index"))
        execute(f"UPDATE leads SET assigned_manager_id=? WHERE id IN ({placeholders})",
                [new_manager_id] + lead_ids)
        log_activity(f"Массовое переназначение {len(lead_ids)} лидов на менеджера {target['name']}.", manager_id)
        flash(f"✅ {len(lead_ids)} лидов переназначено на {target['name']}.", "success")

    else:
        flash("Неизвестное действие.", "error")

    return redirect(url_for("leads.index"))


@bp.route("/leads/<int:lead_id>/participate", methods=["POST"])
@trainer_required
def log_participation(lead_id):
    manager_id, role = session["manager_id"], session["role"]
    lead = query_one("SELECT * FROM leads WHERE id=?", (lead_id,))
    if not lead or not _can_touch(lead, manager_id, role):
        flash("Лид не найден или не ваш.", "error")
        return redirect(url_for("leads.index"))

    round_slot = request.form.get("round_slot", "")
    if round_slot not in ROUND_SLOTS:
        flash("Некорректный раунд розыгрыша.", "error")
        return redirect(url_for("leads.index"))
    round_date = request.form.get("round_date") or date.today().isoformat()

    already = query_one("SELECT 1 FROM participation_log WHERE lead_id=? AND round_date=? AND round_slot=?",
                        (lead_id, round_date, round_slot))
    if already:
        flash("Уже отмечено для этого раунда.", "error")
        return redirect(url_for("leads.index"))

    execute("INSERT INTO participation_log (lead_id, round_date, round_slot) VALUES (?, ?, ?)",
            (lead_id, round_date, round_slot))
    count = query_one("SELECT COUNT(*) c FROM participation_log WHERE lead_id=?", (lead_id,))["c"]
    new_status = "returning" if count >= 2 else "participated"
    execute("UPDATE leads SET participation_count=?, status=?, last_status_change=datetime('now') WHERE id=?",
            (count, new_status, lead_id))
    flash(f"Участие засчитано (всего: {count}).", "success")
    return redirect(url_for("leads.index"))


    return redirect(profile_url)

@bp.route("/leads/<int:lead_id>/adjust-balance", methods=["POST"])
@admin_required
def adjust_balance(lead_id):
    lead = query_one("SELECT * FROM leads WHERE id=?", (lead_id,))
    if not lead:
        flash("Лид не найден.", "error")
        return redirect(url_for("leads.index"))

    try:
        amount = float(request.form.get("amount", 0))
    except ValueError:
        amount = 0
    if amount == 0:
        flash("Укажите сумму (можно отрицательную).", "error")
        return redirect(url_for("leads.index"))
    note = request.form.get("note", "").strip()
    admin_id = session["manager_id"]

    try:
        with db.transaction() as conn:
            cur = conn.execute("UPDATE leads SET balance = balance + ? WHERE id=? AND balance + ? >= 0",
                               (amount, lead_id, amount))
            if cur.rowcount == 0:
                raise ValueError("would_go_negative")
            conn.execute("INSERT INTO balance_ledger (lead_id, amount, reason, actor_manager_id, note) "
                        "VALUES (?, ?, 'admin_adjustment', ?, ?)", (lead_id, amount, admin_id, note))
    except ValueError:
        flash("Итоговый баланс не может стать отрицательным.", "error")
        return redirect(url_for("leads.index"))

    if lead["assigned_manager_id"]:
        from .notifications import notify
        notify(lead["assigned_manager_id"],
              f"Баланс лида {lead['name'] or lead['vk_id']} изменён на {amount:+.2f}G администратором."
              + (f" Причина: {note}" if note else ""),
              url_for("leads.index"))
    log_activity(f"Ручная корректировка баланса лида {lead['name'] or lead['vk_id']}: {amount:+.2f}G "
                f"(было {lead['balance']:.2f}G → стало {lead['balance'] + amount:.2f}G)."
                + (f" Причина: {note}" if note else ""), admin_id)
    flash("Баланс обновлён.", "success")
    return redirect(url_for("leads.index"))


@bp.route("/leads/<int:lead_id>")
@trainer_required
def view(lead_id):
    manager_id = session["manager_id"]
    role = session["role"]
    
    if role == "admin":
        lead = query_one("SELECT * FROM leads WHERE id=?", (lead_id,))
    else:
        lead = query_one("SELECT * FROM leads WHERE id=? AND assigned_manager_id=?", (lead_id, manager_id))
    
    if not lead:
        flash("Лид не найден или не ваш.", "error")
        return redirect(url_for("leads.index"))
    
    scripts = query_all("SELECT id, label FROM scripts WHERE is_active=1 ORDER BY label")
    
    return render_template("lead_detail.html", 
                          lead=dict(lead), 
                          statuses=STATUSES, 
                          status_labels=STATUS_LABELS,
                          scripts=[dict(s) for s in scripts])


# ==================== VK ИНТЕГРАЦИЯ ДЛЯ ЛИДОВ ====================

from .vk_integration import get_direct_message_url, get_vk_profile_url
from flask import redirect, flash, url_for
from .security import login_required
from .db import query_one


@bp.route('/lead/<int:lead_id>/vk-chat')
@trainer_required
def lead_vk_chat(lead_id):
    """Открыть прямой чат с лидом в VK"""
    lead = query_one("SELECT * FROM leads WHERE id = ?", (lead_id,))
    
    if not lead:
        flash('❌ Лид не найден', 'error')
        return redirect(url_for('leads.index'))
    
    if not lead['vk_id']:
        flash('❌ VK ID лида не указан', 'error')
        return redirect(url_for('leads.index'))
    
    chat_url = get_direct_message_url(lead['vk_id'])
    
    if not chat_url:
        flash('❌ Не удалось открыть чат', 'error')
        return redirect(url_for('leads.index'))
    
    return redirect(chat_url)


@bp.route('/lead/<int:lead_id>/vk-profile')
@trainer_required
def lead_vk_profile(lead_id):
    """Открыть профиль лида в VK"""
    lead = query_one("SELECT * FROM leads WHERE id = ?", (lead_id,))
    
    if not lead:
        flash('❌ Лид не найден', 'error')
        return redirect(url_for('leads.index'))
    
    if not lead['vk_id']:
        flash('❌ VK ID лида не указан', 'error')
        return redirect(url_for('leads.index'))
    
    profile_url = get_vk_profile_url(lead['vk_id'])
    
    if not profile_url:
        flash('❌ Не удалось открыть профиль', 'error')
        return redirect(url_for('leads.index'))
    
    return redirect(profile_url)


# ==================== ТЕГИ ====================

@bp.route("/leads/<int:lead_id>/tags/add", methods=["POST"])
@trainer_required
def add_tag(lead_id):
    import json as _json
    manager_id, role = session["manager_id"], session["role"]
    lead = query_one("SELECT * FROM leads WHERE id=?", (lead_id,))
    if not lead or not _can_touch(lead, manager_id, role):
        flash("Лид не найден или не ваш.", "error")
        return redirect(url_for("leads.index"))

    tag = request.form.get("tag", "").strip()[:24]
    if tag:
        tags = _json.loads(lead["tags"] or "[]")
        if tag not in tags:
            tags.append(tag)
            execute("UPDATE leads SET tags=? WHERE id=?", (_json.dumps(tags, ensure_ascii=False), lead_id))
    return redirect(url_for("leads.index"))


@bp.route("/leads/<int:lead_id>/tags/remove", methods=["POST"])
@trainer_required
def remove_tag(lead_id):
    import json as _json
    manager_id, role = session["manager_id"], session["role"]
    lead = query_one("SELECT * FROM leads WHERE id=?", (lead_id,))
    if not lead or not _can_touch(lead, manager_id, role):
        flash("Лид не найден или не ваш.", "error")
        return redirect(url_for("leads.index"))

    tag = request.form.get("tag", "").strip()
    tags = [t for t in _json.loads(lead["tags"] or "[]") if t != tag]
    execute("UPDATE leads SET tags=? WHERE id=?", (_json.dumps(tags, ensure_ascii=False), lead_id))
    return redirect(url_for("leads.index"))


# ==================== ЗАМЕТКИ ====================

@bp.route("/leads/<int:lead_id>/notes/add", methods=["POST"])
@trainer_required
def add_note(lead_id):
    manager_id, role = session["manager_id"], session["role"]
    lead = query_one("SELECT * FROM leads WHERE id=?", (lead_id,))
    if not lead or not _can_touch(lead, manager_id, role):
        flash("Лид не найден или не ваш.", "error")
        return redirect(url_for("leads.index"))

    text = request.form.get("text", "").strip()[:2000]
    if text:
        execute("INSERT INTO lead_notes (lead_id, manager_id, text) VALUES (?, ?, ?)",
                (lead_id, manager_id, text))
        flash("Заметка добавлена.", "success")
    return redirect(url_for("leads.index"))


# ==================== ИМПОРТ ИЗ ФАЙЛА (CSV / Excel) ====================

def _extract_vk_links_from_csv_bytes(data: bytes) -> list:
    """Разбирает CSV: берёт первую колонку каждой строки (или всю строку,
    если запятых нет), пропускает пустые. Кодировка utf-8 с фоллбеком
    на cp1251 (частый случай для файлов, сохранённых из Excel на Windows)."""
    import csv
    import io
    for enc in ("utf-8-sig", "cp1251"):
        try:
            text = data.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = data.decode("utf-8", errors="ignore")
    rows = list(csv.reader(io.StringIO(text)))
    return [row[0].strip() for row in rows if row and row[0].strip()]


def _extract_vk_links_from_xlsx_bytes(data: bytes):
    """Возвращает (список_ссылок, ошибка_или_None). Требует пакет
    openpyxl (`pip install openpyxl`) — опционально, как qrcode для 2FA:
    без него .xlsx не читается, но .csv работает всегда через stdlib."""
    try:
        import io
        import openpyxl
    except ImportError:
        return None, "Для импорта .xlsx нужен пакет openpyxl (pip install openpyxl). Сохрани файл как .csv — он сработает без доп. пакетов."
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    links = []
    for row in ws.iter_rows(values_only=True):
        if row and row[0]:
            links.append(str(row[0]).strip())
    return links, None


@bp.route("/leads/import-file", methods=["POST"])
@admin_required
def import_file():
    file = request.files.get("import_file")
    source_group = request.form.get("source_group", "").strip()
    manager_id = request.form.get("manager_id", type=int) or None

    if not file or not file.filename:
        flash("Файл не выбран.", "error")
        return redirect(url_for("leads.index"))

    data = file.read()
    filename = file.filename.lower()
    if filename.endswith(".xlsx"):
        links, err = _extract_vk_links_from_xlsx_bytes(data)
        if err:
            flash(err, "error")
            return redirect(url_for("leads.index"))
    else:
        links = _extract_vk_links_from_csv_bytes(data)

    if not links:
        flash("В файле не нашлось ни одной ссылки.", "error")
        return redirect(url_for("leads.index"))

    added, duplicates, skipped = bulk_import(links, source_group, manager_id)
    flash(f"📦 Импорт из файла: добавлено {added}, дублей {duplicates}, пропущено {skipped}.", "success")
    return redirect(url_for("leads.index"))

