"""Managers — staff accounts (admin creates/edits), plus per-manager analytics."""
import io
import secrets
import string
from datetime import date, timedelta

from flask import Blueprint, Response, flash, redirect, render_template, request, session, url_for

from . import db
from .db import execute, log_activity, query_all, query_one
from .notifications import notify
from .security import admin_required, hash_password

bp = Blueprint("managers", __name__)


def _gen_password(length=10):
    alphabet = string.ascii_letters + string.digits
    return "".join(secrets.choice(alphabet) for _ in range(length))


def _worked_seconds_expr():
    return (
        "total_seconds_worked + CASE WHEN session_started_at IS NOT NULL "
        "THEN CAST((julianday('now') - julianday(session_started_at)) * 86400 AS INTEGER) "
        "ELSE 0 END"
    )


def _fmt_hours(total_seconds):
    total_seconds = int(total_seconds or 0)
    h, rem = divmod(total_seconds, 3600)
    m = rem // 60
    return f"{h} ч {m} мин"


def manager_stats():
    today = date.today().isoformat()
    week_ago = (date.today() - timedelta(days=7)).isoformat()

    rows = query_all(f"""
        SELECT m.id, m.name, m.login, m.is_active, m.balance, m.total_earned,
          {_worked_seconds_expr()} AS worked_seconds,
          (SELECT COUNT(*) FROM leads WHERE assigned_manager_id=m.id) AS leads_total,
          (SELECT COUNT(*) FROM leads WHERE assigned_manager_id=m.id AND date(found_at)=?) AS leads_today,
          (SELECT COUNT(*) FROM leads WHERE assigned_manager_id=m.id AND date(found_at)>=?) AS leads_week,
          (SELECT COUNT(*) FROM submissions WHERE manager_id=m.id AND status='approved') AS processed_count,
          (SELECT COALESCE(SUM(bl.amount), 0) FROM balance_ledger bl
             JOIN leads l ON l.id = bl.lead_id
             WHERE l.assigned_manager_id = m.id AND bl.reason = 'submission_approved') AS earnings,
          (SELECT COALESCE(SUM(balance), 0) FROM leads WHERE assigned_manager_id=m.id) AS pending_balance
        FROM managers m
        WHERE m.role = 'manager'
        ORDER BY earnings DESC
    """, (today, week_ago))

    stats = []
    for r in rows:
        d = dict(r)
        d["worked_label"] = _fmt_hours(d["worked_seconds"])
        stats.append(d)
    return stats


@bp.route("/managers")
@admin_required
def index():
    stats = manager_stats()
    top3 = sorted(stats, key=lambda x: x["earnings"], reverse=True)[:3]
    return render_template("managers.html", managers=stats, top3=top3)


@bp.route("/managers/create", methods=["POST"])
@admin_required
def create():
    login_name = request.form.get("login", "").strip().lower()
    password = request.form.get("password", "")
    name = request.form.get("name", "").strip()
    if not (login_name and password and name):
        flash("Заполните логин, пароль и имя.", "error")
        return redirect(url_for("managers.index"))
    if len(password) < 6:
        flash("Пароль должен быть не короче 6 символов.", "error")
        return redirect(url_for("managers.index"))
    if query_one("SELECT 1 FROM managers WHERE login=?", (login_name,)):
        flash("Такой логин уже занят.", "error")
        return redirect(url_for("managers.index"))

    execute("INSERT INTO managers (login, password_hash, name, role, total_earned) VALUES (?, ?, ?, 'manager', 0)",
            (login_name, hash_password(password), name))
    log_activity(f"Администратор создал менеджера {name} ({login_name}).")
    flash(f"Менеджер {name} добавлен.", "success")
    return redirect(url_for("managers.index"))


@bp.route("/managers/<int:manager_id>/edit", methods=["POST"])
@admin_required
def edit(manager_id):
    m = query_one("SELECT * FROM managers WHERE id=? AND role='manager'", (manager_id,))
    if not m:
        flash("Менеджер не найден.", "error")
        return redirect(url_for("managers.index"))

    name = request.form.get("name", "").strip()
    new_login = request.form.get("login", "").strip().lower()
    new_password = request.form.get("password", "").strip()

    if not (name and new_login):
        flash("Имя и логин обязательны.", "error")
        return redirect(url_for("managers.index"))
    clash = query_one("SELECT 1 FROM managers WHERE login=? AND id<>?", (new_login, manager_id))
    if clash:
        flash("Такой логин уже занят другим аккаунтом.", "error")
        return redirect(url_for("managers.index"))
    if new_password and len(new_password) < 6:
        flash("Пароль должен быть не короче 6 символов.", "error")
        return redirect(url_for("managers.index"))

    if new_password:
        execute("UPDATE managers SET name=?, login=?, password_hash=? WHERE id=?",
                (name, new_login, hash_password(new_password), manager_id))
        log_activity(f"Администратор изменил логин/пароль менеджера {name} ({new_login}).")
    else:
        execute("UPDATE managers SET name=?, login=? WHERE id=?", (name, new_login, manager_id))
        log_activity(f"Администратор изменил данные менеджера {name} ({new_login}).")

    flash("Данные менеджера обновлены.", "success")
    return redirect(url_for("managers.index"))


@bp.route("/managers/<int:manager_id>/toggle", methods=["POST"])
@admin_required
def toggle(manager_id):
    """Включение/отключение менеджера"""
    m = query_one("SELECT is_active, name FROM managers WHERE id=? AND role='manager'", (manager_id,))
    if not m:
        flash("Менеджер не найден.", "error")
        return redirect(url_for("managers.index"))
    new_status = 0 if m["is_active"] else 1
    execute("UPDATE managers SET is_active = ? WHERE id=?", (new_status, manager_id))
    log_activity(f"Администратор {'отключил' if new_status == 0 else 'включил'} менеджера {m['name']}.")
    flash("Статус обновлён.", "success")
    return redirect(url_for("managers.index"))


@bp.route("/managers/<int:manager_id>/adjust-balance", methods=["POST"])
@admin_required
def adjust_balance(manager_id):
    """Корректировка баланса менеджера"""
    m = query_one("SELECT * FROM managers WHERE id=? AND role='manager'", (manager_id,))
    if not m:
        flash("Менеджер не найден.", "error")
        return redirect(url_for("managers.index"))

    try:
        amount = float(request.form.get("amount", 0))
    except ValueError:
        amount = 0
    if amount == 0:
        flash("Укажите сумму (можно отрицательную).", "error")
        return redirect(url_for("managers.index"))
    note = request.form.get("note", "").strip()
    admin_id = session.get("manager_id")

    try:
        with db.transaction() as conn:
            cur = conn.execute("UPDATE managers SET balance = balance + ? WHERE id=? AND balance + ? >= 0",
                               (amount, manager_id, amount))
            if cur.rowcount == 0:
                raise ValueError("would_go_negative")
            conn.execute("INSERT INTO manager_ledger (manager_id, amount, reason, actor_manager_id, note) "
                        "VALUES (?, ?, 'admin_adjustment', ?, ?)", (manager_id, amount, admin_id, note))
    except ValueError:
        flash("Итоговый баланс не может стать отрицательным.", "error")
        return redirect(url_for("managers.index"))

    notify(manager_id, f"Ваш баланс изменён на {amount:+.2f}G администратором."
          + (f" Причина: {note}" if note else ""), url_for("withdrawals.index"))
    log_activity(f"Ручная корректировка баланса менеджера {m['name']}: {amount:+.2f}G."
                + (f" Причина: {note}" if note else ""), admin_id)
    flash("Баланс обновлён.", "success")
    return redirect(url_for("managers.index"))


@bp.route("/managers/import", methods=["POST"])
@admin_required
def import_excel():
    """Импорт менеджеров из Excel"""
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Прикрепите файл .xlsx.", "error")
        return redirect(url_for("managers.index"))

    try:
        from openpyxl import load_workbook
    except ImportError:
        flash("Импорт из Excel недоступен: не установлен пакет openpyxl.", "error")
        return redirect(url_for("managers.index"))

    try:
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception:
        flash("Не удалось прочитать файл. Проверьте, что это корректный .xlsx.", "error")
        return redirect(url_for("managers.index"))

    if not rows:
        flash("Файл пуст.", "error")
        return redirect(url_for("managers.index"))

    header = [str(c or "").strip().lower() for c in rows[0]]

    def find_col(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None

    col_name = find_col("имя", "name", "фио")
    col_login = find_col("логин", "login")
    col_password = find_col("пароль", "password")

    if col_name is None:
        flash("В файле не найден столбец 'Имя'.", "error")
        return redirect(url_for("managers.index"))

    created, skipped = [], 0
    for row in rows[1:]:
        if not row or col_name >= len(row):
            continue
        name = str(row[col_name] or "").strip()
        if not name:
            continue

        login_name = str(row[col_login]).strip().lower() if col_login is not None and row[col_login] else ""
        if not login_name:
            base = "".join(ch for ch in name.lower().replace(" ", ".") if ch.isalnum() or ch == ".")
            login_name = base or f"manager{secrets.randbelow(9999)}"
        candidate, n = login_name, 1
        while query_one("SELECT 1 FROM managers WHERE login=?", (candidate,)):
            n += 1
            candidate = f"{login_name}{n}"
        login_name = candidate

        password = str(row[col_password]).strip() if col_password is not None and row[col_password] else ""
        if not password or len(password) < 6:
            password = _gen_password()

        execute("INSERT INTO managers (login, password_hash, name, role) VALUES (?, ?, ?, 'manager')",
                (login_name, hash_password(password), name))
        created.append((name, login_name, password))

    if not created:
        flash("Не найдено ни одной строки для импорта.", "error")
        return redirect(url_for("managers.index"))

    log_activity(f"Администратор импортировал {len(created)} менеджеров из Excel.")

    lines = ["Имя\tЛогин\tПароль"] + [f"{n}\t{l}\t{p}" for n, l, p in created]
    buf = io.BytesIO("\r\n".join(lines).encode("utf-8-sig"))
    flash(f"Импортировано менеджеров: {len(created)}. Скачивается файл с логинами и паролями.", "success")
    return Response(
        buf.getvalue(), mimetype="text/plain",
        headers={"Content-Disposition": "attachment; filename=managers_import_credentials.txt"})


@bp.route("/managers/export")
@admin_required
def export_txt():
    """Экспорт логинов в TXT"""
    rows = query_all("SELECT name, login, is_active FROM managers WHERE role='manager' ORDER BY name")
    lines = ["Имя\tЛогин\tСтатус"]
    for r in rows:
        lines.append(f"{r['name']}\t{r['login']}\t{'активен' if r['is_active'] else 'отключён'}")
    buf = io.BytesIO("\r\n".join(lines).encode("utf-8-sig"))
    log_activity("Администратор выгрузил список логинов менеджеров в TXT.")
    return Response(
        buf.getvalue(), mimetype="text/plain",
        headers={"Content-Disposition": "attachment; filename=managers_logins.txt"})